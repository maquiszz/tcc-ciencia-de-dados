import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

print("Iniciando programa...")

# Janela invisível
root = tk.Tk()
root.withdraw()
root.attributes('-topmost', True)

# Selecionar arquivo
arquivo_csv = filedialog.askopenfilename(
    title="Selecione um arquivo CSV",
    filetypes=[("Arquivos CSV", "*.csv")]
)

if arquivo_csv:

    print(f"Arquivo selecionado: {arquivo_csv}")

    # Leitura robusta
    df = pd.read_csv(
        arquivo_csv,
        sep=None,
        engine='python',
        encoding='latin1',
        on_bad_lines='skip'
    )

    # Nome do arquivo Excel
    arquivo_excel = arquivo_csv.replace(".csv", ".xlsx")

    # Converter para Excel
    df.to_excel(arquivo_excel, index=False)

    # Abrir workbook
    wb = load_workbook(arquivo_excel)
    ws = wb.active

    # ==========================
    # ESTILOS
    # ==========================

    # Cabeçalho
    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    thin = Side(border_style="thin", color="D3D3D3")

    # Aplicar estilo no cabeçalho
    for cell in ws[1]:

        cell.fill = header_fill
        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        cell.border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

    # ==========================
    # AJUSTAR TAMANHO COLUNAS
    # ==========================

    for column in ws.columns:

        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:

            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = max_length + 4

        ws.column_dimensions[column_letter].width = adjusted_width

    # ==========================
    # FORMATAR CÉLULAS
    # ==========================

    for row in ws.iter_rows():

        for cell in row:

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            cell.border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin
            )

    # ==========================
    # FILTRO AUTOMÁTICO
    # ==========================

    ws.auto_filter.ref = ws.dimensions

    # ==========================
    # CONGELAR CABEÇALHO
    # ==========================

    ws.freeze_panes = "A2"

    # ==========================
    # SALVAR
    # ==========================

    wb.save(arquivo_excel)

    print("\nARQUIVO CONVERTIDO E FORMATADO COM SUCESSO!")
    print(f"Arquivo salvo em:\n{arquivo_excel}")

else:
    print("Nenhum arquivo selecionado.")
